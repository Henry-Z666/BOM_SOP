
import os
from openpyxl import load_workbook

root = r'C:\Users\10602\Desktop\AI_assembly'
src_path = os.path.join(root, 'SOP示例.xlsx')
ref_path = os.path.join(root, 'data', 'runs', 'reference_template.xlsx')
os.makedirs(os.path.dirname(ref_path), exist_ok=True)

wb = load_workbook(src_path)
src_ws = wb['补液箱分装-1']

step_names = [
    '第1步-固定水箱焊件',
    '第2步-安装顶板焊件',
    '第3步-安装软管',
    '第4步-安装卡式端盖',
    '第5步-安装压力传感器',
    '第6步-安装球阀和转接头',
    '第7步-安装电磁阀',
    '第8步-安装进水管焊件',
]

for name in step_names:
    ws = wb.copy_worksheet(src_ws)
    ws.title = name

for name in ['对应力矩一览表', '补液箱分装-1', '补液箱分装-2', '补液箱分装-3', '水箱检漏1', '水箱检漏 2']:
    try:
        del wb[name]
    except:
        pass

wb.save(ref_path)
print(f'Saved: {ref_path}')
print(f'Sheets: {wb.sheetnames}')
