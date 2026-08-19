from __future__ import annotations

from pathlib import Path
import tempfile
import unittest

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
from PIL import Image

from sop_pipeline.agent.sop_publisher import (
    SopImage,
    SopStep,
    SopPublisher,
    _column_width_pixels,
    _row_height_pixels,
)


def _image(folder: Path, name: str) -> Path:
    path = folder / name
    Image.new("RGB", (320, 240), "white").save(path)
    return path


def _step(index: int, process: int, image: Path, **changes) -> SopStep:
    values = dict(
        step_id=f"step-{index}",
        main_process_id=str(process),
        main_process_name=f"主工序 {process}",
        title=f"安装步骤 {index}",
        image=SopImage(f"image-{index}", image),
        materials=((f"M-{index}", "零件", 1),),
        process_text="按图装配",
        control_points="确认安装到位",
        tools="常规工具",
    )
    values.update(changes)
    return SopStep(**values)


class SopPublisherTests(unittest.TestCase):
    def test_42_images_are_grouped_by_8_processes_with_continuation_pages(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            steps = []
            index = 1
            for process in range(1, 9):
                count = 6 if process <= 2 else 5
                for _ in range(count):
                    steps.append(_step(index, process, _image(root, f"{index}.jpg")))
                    index += 1
            workbook_path = SopPublisher(images_per_page=2).publish(
                tuple(steps), root / "交付结果"
            )
            workbook = load_workbook(workbook_path)

            self.assertEqual(sum(len(sheet._images) for sheet in workbook), 42)
            self.assertTrue(any("续页" in name for name in workbook.sheetnames))
            self.assertEqual(len(workbook.sheetnames), 24)
            self.assertEqual(
                {path.name for path in (root / "交付结果").iterdir()},
                {"SOP.xlsx", "步骤图片"},
            )

    def test_100_main_processes_do_not_assume_two_digit_sheet_numbers(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            shared_image = _image(root, "shared.jpg")
            steps = tuple(
                _step(index, index, shared_image) for index in range(1, 101)
            )
            workbook_path = SopPublisher().publish(steps, root / "交付结果")
            workbook = load_workbook(workbook_path, read_only=False)

        self.assertEqual(len(workbook.sheetnames), 100)
        self.assertIn("主工序 100", workbook.sheetnames)

    def test_pending_delivery_keeps_candidates_then_final_publish_cleans_them(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            recommended = _image(root, "recommended.jpg")
            alternate = _image(root, "alternate.jpg")
            step = _step(
                1,
                1,
                recommended,
                questioned=True,
                candidates=(
                    SopImage("candidate-a", recommended, "a", True),
                    SopImage("candidate-b", alternate, "b", False),
                ),
            )
            publisher = SopPublisher()
            pending = publisher.publish((step,), root / "交付结果", pending=True)
            pending_book = load_workbook(pending)

            self.assertEqual(pending.name, "SOP_待确认.xlsx")
            self.assertIn("待确认", pending_book.active["AN5"].value)
            self.assertEqual(len(list((root / "交付结果" / "步骤图片").iterdir())), 2)

            final_step = _step(1, 1, recommended)
            final = publisher.publish((final_step,), root / "交付结果", pending=False)

            self.assertEqual(final.name, "SOP.xlsx")
            self.assertFalse((root / "交付结果" / "SOP_待确认.xlsx").exists())
            self.assertEqual(len(list((root / "交付结果" / "步骤图片").iterdir())), 1)

    def test_publication_fills_the_retained_single_page_template(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            image = _image(root, "step.jpg")
            workbook_path = SopPublisher().publish(
                (_step(1, 1, image, control_points="", tools=""),),
                root / "交付结果",
            )
            workbook = load_workbook(workbook_path)
            sheet = workbook.active

            self.assertEqual(sheet["B7"].value, "装配内容")
            self.assertEqual(sheet["AI19"].value, "物  料  表")
            self.assertEqual(sheet["AI30"].value, "工  具/工  装  表")
            self.assertEqual(sheet["AN4"].value, "1")
            self.assertEqual(sheet["AN5"].value, "主工序 1")
            self.assertEqual(sheet["AJ8"].value, "按图装配")
            self.assertEqual(sheet["AJ32"].value, "待填写")
            self.assertEqual(len(sheet._images), 1)
            self.assertIn("AM21:AQ21", {str(item) for item in sheet.merged_cells.ranges})

    def test_publication_removes_stale_external_defined_names(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            image = _image(root, "step.jpg")
            source_template = load_workbook(
                Path(__file__).parents[1] / "assets" / "sop-template.xlsx"
            )
            source_template.defined_names.add(
                DefinedName("stale_external_name", attr_text="[1]目录!#REF!")
            )
            stale_template = root / "stale-template.xlsx"
            source_template.save(stale_template)
            source_template.close()

            workbook_path = SopPublisher(template_path=stale_template).publish(
                (_step(1, 1, image),),
                root / "交付结果",
            )
            workbook = load_workbook(workbook_path)

            stale_names = [
                name
                for name, definition in workbook.defined_names.items()
                if "#REF!" in str(definition.attr_text)
                or str(definition.attr_text).startswith("[")
            ]
            self.assertEqual(stale_names, [])

    def test_one_main_process_places_its_installation_images_on_one_page(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            steps = tuple(
                _step(index, 1, _image(root, f"{index}.jpg"))
                for index in range(1, 7)
            )
            workbook_path = SopPublisher().publish(steps, root / "交付结果")
            workbook = load_workbook(workbook_path)

            self.assertEqual(workbook.sheetnames, ["主工序 1"])
            self.assertEqual(len(workbook.active._images), 6)
            self.assertEqual(workbook.active["AN5"].value, "主工序 1")
            self.assertEqual(workbook.active["AJ8"].value, "按图装配")
            for image in workbook.active._images:
                marker = image.anchor._from
                self.assertLess(
                    marker.colOff,
                    _column_width_pixels(workbook.active, marker.col + 1) * 9525,
                )
                self.assertLess(
                    marker.rowOff,
                    _row_height_pixels(workbook.active, marker.row + 1) * 9525,
                )

    def test_a_main_process_over_template_capacity_uses_continuation_page(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            steps = tuple(
                _step(index, 1, _image(root, f"{index}.jpg"))
                for index in range(1, 8)
            )
            workbook_path = SopPublisher().publish(steps, root / "交付结果")
            workbook = load_workbook(workbook_path)

            self.assertEqual(workbook.sheetnames, ["主工序 1", "主工序 1-续页2"])
            self.assertEqual([len(sheet._images) for sheet in workbook], [4, 3])


if __name__ == "__main__":
    unittest.main()
