"""Validate the grouped, publishable eight-process SOP workbook."""
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[3]
OUTPUT = ROOT / "outputs" / "published_sop" / "JB9918900337_水箱部件装配SOP_出版版.xlsx"

EXPECTED = {
    "第1步-固定水箱焊件": 13,
    "第2步-安装顶板焊件": 4,
    "第3步-安装软管": 3,
    "第4步-安装卡式端盖": 3,
    "第5步-安装压力传感器": 2,
    "第6步-安装球阀和转接头": 4,
    "第7步-安装电磁阀": 6,
    "第8步-安装进水管焊件": 7,
}

if not OUTPUT.exists():
    raise SystemExit(f"missing workbook: {OUTPUT}")

wb = load_workbook(OUTPUT, data_only=False)
errors: list[str] = []
if wb.sheetnames != list(EXPECTED):
    errors.append(f"sheet order mismatch: {wb.sheetnames}")

total_images = 0
for index, (name, expected_images) in enumerate(EXPECTED.items(), 1):
    if name not in wb.sheetnames:
        continue
    ws = wb[name]
    total_images += len(ws._images)
    if len(ws._images) != expected_images:
        errors.append(f"{name}: expected {expected_images} images, got {len(ws._images)}")
    if ws["AB4"].value != f"FZ.1-{index}":
        errors.append(f"{name}: wrong process code {ws['AB4'].value!r}")
    if ws["AB5"].value not in name:
        errors.append(f"{name}: wrong process name {ws['AB5'].value!r}")
    if "JB9918900337-SOP-001" not in str(ws["B4"].value):
        errors.append(f"{name}: missing file number")
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and any(token in cell.value for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")):
                errors.append(f"{name}!{cell.coordinate}: formula error {cell.value}")

if total_images != 42:
    errors.append(f"expected 42 total images, got {total_images}")

if errors:
    print("PUBLISHED SOP VALIDATION FAILED")
    for error in errors:
        print("-", error)
    raise SystemExit(1)

print("PUBLISHED SOP VALIDATION PASSED")
print(f"workbook={OUTPUT}")
print(f"sheets={len(wb.sheetnames)} images={total_images}")
for name, count in EXPECTED.items():
    print(f"  {name}: {count}")
