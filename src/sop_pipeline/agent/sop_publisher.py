from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from itertools import groupby
from math import floor
from pathlib import Path
import re
import shutil
from typing import Protocol

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as WorksheetImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, points_to_pixels

from .bundle_paths import bundled_sop_template
from .sop_layout import PixelSize, balanced_page_sizes, plan_page_layout


@dataclass(frozen=True)
class SopImage:
    image_id: str
    path: Path
    candidate_id: str | None = None
    recommended: bool = False
    placeholder: bool = False


@dataclass(frozen=True)
class SopStep:
    step_id: str
    main_process_id: str
    main_process_name: str
    title: str
    image: SopImage
    materials: tuple[tuple[str, str, int], ...]
    process_text: str
    control_points: str
    tools: str
    project_name: str = "待填写"
    document_no: str = "待填写"
    applicable_model: str = "待填写"
    applicable_base: str = "待填写"
    questioned: bool = False
    candidates: tuple[SopImage, ...] = ()


class WorkbookVerifier(Protocol):
    def verify(self, workbook_path: Path) -> None: ...


class OpenpyxlWorkbookVerifier:
    def verify(self, workbook_path: Path) -> None:
        workbook = load_workbook(workbook_path, read_only=False, data_only=False)
        if not workbook.sheetnames:
            raise ValueError("published workbook has no worksheets")
        for sheet in workbook.worksheets:
            if not sheet.print_area:
                raise ValueError(f"worksheet has no print area: {sheet.title}")
            if not sheet._images:
                raise ValueError(f"worksheet has no step image: {sheet.title}")
        workbook.close()


class SopPublisher:
    """Built-in, main-process-paginated XLSX publisher with a strict delivery root."""

    def __init__(
        self,
        *,
        images_per_page: int = 6,
        verifier: WorkbookVerifier | None = None,
        template_path: Path | None = None,
    ) -> None:
        if not 1 <= images_per_page <= 6:
            raise ValueError("images_per_page must be between 1 and 6")
        self.images_per_page = images_per_page
        self.verifier = verifier or OpenpyxlWorkbookVerifier()
        self.template_path = Path(template_path) if template_path else bundled_sop_template()

    def publish(
        self,
        steps: tuple[SopStep, ...],
        delivery_directory: Path,
        *,
        pending: bool = False,
    ) -> Path:
        if not steps:
            raise ValueError("SOP publication requires at least one step")
        delivery_directory.mkdir(parents=True, exist_ok=True)
        image_directory = delivery_directory / "步骤图片"
        image_directory.mkdir(parents=True, exist_ok=True)

        copied = self._copy_delivery_images(steps, image_directory, pending=pending)
        self._clean_known_delivery_artifacts(
            delivery_directory,
            image_directory,
            keep_images=set(copied.values()),
            pending=pending,
        )
        workbook = self._build_workbook(
            steps, copied, image_directory, pending=pending
        )
        filename = "SOP_待确认.xlsx" if pending else "SOP.xlsx"
        target = delivery_directory / filename
        temporary = delivery_directory / f".{filename}.tmp.xlsx"
        try:
            workbook.save(temporary)
        finally:
            # openpyxl keeps image streams on the workbook until it is closed.
            # Explicit closure is required for long Agent runs and for Windows
            # delivery directories that may be replaced immediately afterward.
            workbook.close()
        self.verifier.verify(temporary)
        temporary.replace(target)
        self._verify_delivery_whitelist(delivery_directory, pending=pending)
        return target

    def _copy_delivery_images(
        self,
        steps: tuple[SopStep, ...],
        image_directory: Path,
        *,
        pending: bool,
    ) -> dict[str, str]:
        copied: dict[str, str] = {}
        ordinal = 1
        for step in steps:
            images = step.candidates if pending and step.candidates else (step.image,)
            for image in images:
                if not image.path.is_file():
                    raise FileNotFoundError(image.path)
                suffix = image.path.suffix.lower() or ".png"
                candidate = f"-{image.candidate_id}" if image.candidate_id else ""
                destination_name = (
                    f"{ordinal:04d}-{_safe_filename(step.step_id)}{candidate}{suffix}"
                )
                destination = image_directory / destination_name
                shutil.copy2(image.path, destination)
                copied[image.image_id] = destination_name
                ordinal += 1
        return copied

    def _build_workbook(
        self,
        steps: tuple[SopStep, ...],
        copied: dict[str, str],
        image_directory: Path,
        *,
        pending: bool,
    ) -> Workbook:
        workbook = load_workbook(self.template_path, read_only=False, data_only=False)
        _drop_stale_external_defined_names(workbook)
        template_sheet = workbook.active
        if template_sheet["B7"].value != "装配内容":
            workbook.close()
            raise ValueError("内置SOP模板结构不匹配：缺少装配内容区域")
        pages: list[tuple[SopStep, ...]] = []
        for _process_id, process_steps in groupby(
            steps, key=lambda item: item.main_process_id
        ):
            grouped = tuple(process_steps)
            offset = 0
            for page_size in balanced_page_sizes(len(grouped), self.images_per_page):
                pages.append(grouped[offset : offset + page_size])
                offset += page_size

        # Copy the still-pristine retained template before filling any page. This
        # prevents values from the first process leaking into later pages.
        sheets = [template_sheet]
        sheets.extend(workbook.copy_worksheet(template_sheet) for _ in pages[1:])
        used_names: set[str] = set()
        process_pages: dict[str, int] = {}
        for sheet, page_steps in zip(sheets, pages, strict=True):
            representative = page_steps[0]
            page = process_pages.get(representative.main_process_id, 0) + 1
            process_pages[representative.main_process_id] = page
            suffix = "" if page == 1 else f"-续页{page}"
            requested = (
                representative.main_process_name or representative.main_process_id
            ) + suffix
            sheet.title = _unique_sheet_name(requested, used_names)
            self._fill_template_page(
                sheet,
                page_steps,
                copied,
                image_directory,
                pending=pending,
            )
        return workbook

    def _fill_template_page(
        self,
        sheet,
        steps: tuple[SopStep, ...],
        copied: dict[str, str],
        image_directory: Path,
        *,
        pending: bool,
    ) -> None:
        """Fill one main-process page in the retained template."""

        step = steps[0]
        sheet.sheet_view.showGridLines = False
        sheet["J1"] = f"项目名称：{_value_or_pending(step.project_name)}"
        sheet["B4"] = f"文件编号：{_value_or_pending(step.document_no)}"
        sheet["M4"] = _value_or_pending(step.applicable_model)
        sheet["V4"] = _value_or_pending(step.applicable_base)
        sheet["AN4"] = str(step.main_process_id)
        process_name = _value_or_pending(step.main_process_name or step.title)
        questioned = any(item.questioned for item in steps)
        sheet["AN5"] = process_name + ("（待确认）" if pending and questioned else "")

        controls = _page_controls(steps)
        for offset, row in enumerate(range(8, 19)):
            sheet[f"AJ{row}"] = controls[offset] if offset < len(controls) else None
        if not controls:
            sheet["AJ8"] = "待填写"

        for row in range(21, 30):
            sheet[f"AJ{row}"] = None
            sheet[f"AM{row}"] = None
            sheet[f"AR{row}"] = None
        materials = _page_materials(steps)
        for row, material in zip(range(21, 30), materials, strict=False):
            code, name, quantity = material
            sheet[f"AJ{row}"] = _value_or_pending(code)
            sheet[f"AM{row}"] = _value_or_pending(name)
            sheet[f"AR{row}"] = quantity if quantity not in (None, "") else "待填写"
        if not materials:
            sheet["AJ21"] = "待填写"
            sheet["AM21"] = "待填写"
            sheet["AR21"] = "待填写"

        tools = _split_entries(*(item.tools for item in steps))
        for row in range(32, 36):
            sheet[f"AJ{row}"] = None
            sheet[f"AM{row}"] = None
            sheet[f"AR{row}"] = None
        for row, tool in zip(range(32, 36), tools, strict=False):
            sheet[f"AJ{row}"] = tool
            sheet[f"AM{row}"] = "待填写"
            sheet[f"AR{row}"] = "待填写"
        if not tools:
            sheet["AJ32"] = "待填写"
            sheet["AM32"] = "待填写"
            sheet["AR32"] = "待填写"

        worksheet_images: list[WorksheetImage] = []
        for item in steps:
            publication_image = _publication_image(item, pending)
            image_bytes = (
                image_directory / copied[publication_image.image_id]
            ).read_bytes()
            worksheet_images.append(WorksheetImage(BytesIO(image_bytes)))
        placements = plan_page_layout(
            tuple(
                PixelSize(int(image.width), int(image.height))
                for image in worksheet_images
            ),
            zone_width=_image_zone_width(sheet),
            zone_height=_image_zone_height(sheet),
        )
        for worksheet_image, placement in zip(
            worksheet_images, placements, strict=True
        ):
            worksheet_image.width = placement.width
            worksheet_image.height = placement.height
            marker = _image_zone_marker(sheet, placement.x, placement.y)
            worksheet_image.anchor = OneCellAnchor(
                _from=marker,
                ext=XDRPositiveSize2D(
                    cx=pixels_to_EMU(placement.width),
                    cy=pixels_to_EMU(placement.height),
                ),
            )
            sheet.add_image(worksheet_image)

        sheet.print_area = "B1:AR35"
        sheet.print_options.horizontalCentered = True
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1
        sheet.sheet_properties.pageSetUpPr.fitToPage = True

    def _clean_known_delivery_artifacts(
        self,
        delivery_directory: Path,
        image_directory: Path,
        *,
        keep_images: set[str],
        pending: bool,
    ) -> None:
        for existing in image_directory.iterdir():
            if existing.is_file() and existing.name not in keep_images:
                existing.unlink()
        obsolete = delivery_directory / ("SOP.xlsx" if pending else "SOP_待确认.xlsx")
        if obsolete.exists():
            obsolete.unlink()

    @staticmethod
    def _verify_delivery_whitelist(delivery_directory: Path, *, pending: bool) -> None:
        workbook_name = "SOP_待确认.xlsx" if pending else "SOP.xlsx"
        actual = {entry.name for entry in delivery_directory.iterdir()}
        expected = {workbook_name, "步骤图片"}
        if actual != expected:
            raise ValueError(f"delivery directory violates whitelist: {sorted(actual)}")


def _value_or_pending(value: object) -> object:
    if value is None:
        return "待填写"
    if isinstance(value, str):
        return value.strip() or "待填写"
    return value


_IMAGE_ZONE_MIN_COLUMN = 2  # B
_IMAGE_ZONE_MAX_COLUMN = 34  # AH
_IMAGE_ZONE_MIN_ROW = 8
_IMAGE_ZONE_MAX_ROW = 35
_EXCEL_MAX_DIGIT_WIDTH = 7
_DEFAULT_RENDERED_ROW_HEIGHT_POINTS = 15.0


def _image_zone_width(sheet) -> int:
    return sum(
        _column_width_pixels(sheet, column)
        for column in range(_IMAGE_ZONE_MIN_COLUMN, _IMAGE_ZONE_MAX_COLUMN + 1)
    )


def _image_zone_height(sheet) -> int:
    return sum(
        _row_height_pixels(sheet, row)
        for row in range(_IMAGE_ZONE_MIN_ROW, _IMAGE_ZONE_MAX_ROW + 1)
    )


def _image_zone_marker(sheet, x: int, y: int) -> AnchorMarker:
    column, column_offset = _axis_marker(
        x,
        _IMAGE_ZONE_MIN_COLUMN,
        _IMAGE_ZONE_MAX_COLUMN,
        lambda index: _column_width_pixels(sheet, index),
    )
    row, row_offset = _axis_marker(
        y,
        _IMAGE_ZONE_MIN_ROW,
        _IMAGE_ZONE_MAX_ROW,
        lambda index: _row_height_pixels(sheet, index),
    )
    return AnchorMarker(
        col=column - 1,
        row=row - 1,
        colOff=pixels_to_EMU(column_offset),
        rowOff=pixels_to_EMU(row_offset),
    )


def _axis_marker(offset: int, start: int, end: int, span) -> tuple[int, int]:
    if offset < 0:
        raise ValueError("image offset cannot be negative")
    remaining = offset
    for index in range(start, end + 1):
        width = span(index)
        if remaining < width:
            return index, remaining
        remaining -= width
    raise ValueError("image offset escapes the SOP image zone")


def _column_width_pixels(sheet, column: int) -> int:
    width = sheet.sheet_format.defaultColWidth or 8.43
    for dimension in sheet.column_dimensions.values():
        if dimension.min <= column <= dimension.max:
            width = dimension.width
            break
    # Excel column widths exclude cell padding. Four pixels matches the retained
    # template's rendered B:AH width (980 px versus 979.3 px measured by Excel).
    character_pixels = floor(
        (
            256 * width
            + floor(128 / _EXCEL_MAX_DIGIT_WIDTH)
        )
        / 256
        * _EXCEL_MAX_DIGIT_WIDTH
    )
    return max(1, character_pixels + 4)


def _row_height_pixels(sheet, row: int) -> int:
    # The retained workbook declares a 20.1 pt default but Excel renders its
    # unspecified rows at 15 pt. Explicit row heights remain authoritative.
    points = sheet.row_dimensions[row].height
    if points is None:
        points = _DEFAULT_RENDERED_ROW_HEIGHT_POINTS
    return max(1, points_to_pixels(points))


def _drop_stale_external_defined_names(workbook: Workbook) -> None:
    """Remove broken links retained by old templates before Excel COM opens them."""

    for name, definition in list(workbook.defined_names.items()):
        target = str(getattr(definition, "attr_text", "") or "").strip()
        if "#REF!" in target or re.match(r"^\[\d+\]", target):
            del workbook.defined_names[name]


def _split_entries(*values: str) -> tuple[str, ...]:
    entries: list[str] = []
    for value in values:
        for item in re.split(r"[\r\n；;]+", str(value or "")):
            normalized = item.strip()
            if normalized and normalized not in entries:
                entries.append(normalized)
    return tuple(entries)


def _page_controls(steps: tuple[SopStep, ...]) -> tuple[str, ...]:
    entries: list[str] = []
    for step in steps:
        descriptions = _split_entries(step.process_text, step.control_points)
        for description in descriptions:
            if description not in entries:
                entries.append(description)
        if not descriptions:
            title = str(step.title or "").strip()
            if title and title not in entries:
                entries.append(title)
    return tuple(entries)


def _page_materials(
    steps: tuple[SopStep, ...],
) -> tuple[tuple[str, str, int], ...]:
    ordered: list[tuple[str, str, int]] = []
    positions: dict[tuple[str, str], int] = {}
    for step in steps:
        for code, name, quantity in step.materials:
            key = (str(code), str(name))
            if key in positions:
                index = positions[key]
                previous = ordered[index]
                ordered[index] = (previous[0], previous[1], previous[2] + quantity)
            else:
                positions[key] = len(ordered)
                ordered.append((code, name, quantity))
    return tuple(ordered)


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r"[<>:\"/\\|?*]", "_", value).strip(" .")
    return cleaned or "step"


def _publication_image(step: SopStep, pending: bool) -> SopImage:
    if pending and step.candidates:
        return next(
            (candidate for candidate in step.candidates if candidate.recommended),
            step.candidates[0],
        )
    return step.image


def _unique_sheet_name(requested: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", requested).strip() or "主工序"
    base = base[:31]
    candidate = base
    ordinal = 2
    while candidate.casefold() in used:
        suffix = f"-{ordinal}"
        candidate = base[: 31 - len(suffix)] + suffix
        ordinal += 1
    used.add(candidate.casefold())
    return candidate
