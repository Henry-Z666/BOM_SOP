"""Clean-run planner v2 (BOM-authoritative rewrite).

Reads the Excel BOM and the discovery CAD graph, then produces:
  1. an authoritative manifest (assembly identity + hash; the calibrated
     cameras are added later by calibrate_cameras.py)
  2. a step plan for the BOM assembly steps

Planning rules implemented here (interpretation of rules/*.md):

  * bom-steps/v2 (portable BOM understanding layer)
      The sheet and every column are discovered from the header row
      (never hardcoded: other equipment BOMs keep the header family but
      differ in sheet name / column order and carry 100+ steps).  Main
      assembly steps come from the process column's "第N步…" anchors;
      render sub-steps are the BOM hierarchy units inside each main
      step, in strict BOM row order.

  * bom-order-authority/v1
      Steps follow the BOM row order exactly.  No geometry-driven
      reordering: moving a part ahead of its receiver produced floating
      installs; executing the BOM order never can.

  * no-silent-skip/v1
      Every BOM row must resolve to CAD occurrences, or be an explicit
      raw-material row (unit kg / metre).  Any other unresolved row
      aborts the plan with a diagnostic; rows whose parts already ride
      inside an ancestor installed by an earlier step are recorded as
      explicit component-detail merges, never dropped silently.

  * cad-match/v1 (multi-key)
      A row is matched against the CAD graph occurrence model stems by
      BOM drawing number first (col 5), then by BOM model/spec (col 7).
      Keys are normalised: case-insensitive, '.' treated as '_'.
      Occurrence pools are allocated to rows in BOM order, quantity-wise.

  * first-placement/v2
      The FIRST main step whose anchor is a single assembly with its
      children listed in the BOM is the base structure built on the
      fixture: the anchor is a container (no own occurrence) and its
      direct children are the sub-steps, the first of which lands as
      one rigid unit with no receiver (vertical placement direction).
      Every other main step installs its direct-child groups as units.

  * explosion-contact/v5
      Receivers are the LEAF occurrences installed so far (root-level
      boxes are far too coarse: a hose inside the tank hull interpenetrates
      every axis and loses every contact).  Insertion mates (shallow
      interpenetration, cap scaled by the receiver extent along the axis)
      outrank touch contacts, because a fastener sliding into a hole
      interpenetrates its receiver while merely brushing a side wall.
      Fallback = shortest hull-clearance travel.  Rotation untouched.
      v5 distance: the 60 mm lower floor used to fling 16 mm O-rings to
      3.7x their own size, reading as stray markers.  The floor now scales
      with the part: min(60, max(1.2*size, insertion_depth+0.5*extent+5)),
      so small parts stay hugging their seat while large parts keep the
      legacy behaviour (their floor saturates at 60 anyway).

  * camera-approach/v3
      Of the two locked views, choose the camera whose eye lies on the
      side the part comes from (view-from dot explosion normal), i.e.
      behind the moving part, so part and receiver face are both visible;
      tie-broken by which camera sees the moving part in front of the
      receiver.  Before camera calibration exists, a documented analytic
      fallback (the two views are vertical mirrors) decides on the Y
      component only.

  * naming/v2
      step_id = "<main:02d>.<sub:02d>_<task_code>" (zero widths grow
      with 100+ step BOMs); bom_level stays in the step for traceability.
      Each plan owns a batch image folder (out/images/<task_code>_
      <timestamp>) so generations never mix.  --reuse-batch keeps the
      existing folder on replans.

usage: python plan_steps.py <task_code> [limit] [--reuse-batch]
"""
from __future__ import annotations

import datetime
import hashlib
import json
import math
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
BOM_FILE = ROOT / "inputs" / "BOM.xlsx"
GRAPH_FILE = ROOT / "out" / "data" / "cad-graph.json"
MODELS_DIR = ROOT / "inputs" / "models"
# pipeline-stability/v1: every tuned constant lives in ONE file; planner
# and Renderer read the same authority so a fix can never desync sides.
CONFIG_FILE = ROOT / "rendering_config.json"
_CFG = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
ASM_FILE = _CFG["assembly_file"]
EXP = _CFG["explosion"]
OUT_MANIFEST = ROOT / "out" / "data" / "manifest.json"
OUT_PLAN = ROOT / "out" / "data" / "plan.json"

# raw-material rows (stock bought by mass/length) never own a CAD
# occurrence; they are consumed by their fabricated parent group.
RAW_UNITS = {"千克", "米"}

_CN_DIG = {"零": 0, "一": 1, "二": 2, "三": 3, "四": 4,
           "五": 5, "六": 6, "七": 7, "八": 8, "九": 9}


def _cn_int(tok: str) -> int:
    """minimal Chinese numeral support (1..999) for step anchors."""
    if not tok:
        return 0
    if "百" in tok:
        a, _, b = tok.partition("百")
        return (_CN_DIG.get(a, 1) or 1) * 100 + _cn_int(b)
    if "十" in tok:
        a, _, b = tok.partition("十")
        tens = _CN_DIG.get(a, 1) if a else 1
        return tens * 10 + (_CN_DIG.get(b, 0) if b else 0)
    return _CN_DIG.get(tok, 0)


def parse_step_no(text: str):
    """'第12步…' / '第十二步…' -> int step number; None if the cell is
    not a main-step anchor."""
    m = re.search(r"第\s*([0-9]+|[一二三四五六七八九十百两]+)\s*步",
                  text or "")
    if not m:
        return None
    tok = m.group(1)
    return int(tok) if tok.isdigit() else _cn_int(tok)


def open_bom_sheet():
    """bom-steps/v2: pick the first worksheet whose header band (rows
    1-3, headers are merged across rows) carries both a 层级 column and
    a process (步) column - sheet names differ across equipment BOMs
    and must never be hardcoded."""
    wb = openpyxl.load_workbook(BOM_FILE, data_only=True)
    for ws in wb.worksheets:
        hdrs = [[str(c.value or "") for c in ws[r]]
                for r in range(1, min(3, ws.max_row) + 1)]
        flat = [h for row in hdrs for h in row]
        if any(("层级" in h or "层次" in h) for h in flat) \
                and any("步" in h for h in flat):
            return ws, hdrs
    raise SystemExit("[PLAN] bom-steps/v2 violated: no worksheet with "
                     "层级/层次 + 步 header columns found in "
                     f"{BOM_FILE}")


def detect_columns(hdrs: list[list[str]]) -> dict[str, int]:
    """header-driven column discovery (1-based); abort loudly on a
    missing family member instead of misreading shifted columns."""
    def find(*keys):
        for row in hdrs:
            for i, h in enumerate(row):
                if any(k in h for k in keys):
                    return i + 1
        return None
    cols = {"level": find("层级", "层次"), "drawing": find("图号"),
            "name": find("名称"), "spec": find("型号"),
            "qty": find("数量"), "unit": find("单位"),
            "step": find("装配", "步")}
    missing = [k for k, v in cols.items() if not v]
    if missing:
        raise SystemExit(f"[PLAN] BOM header missing columns {missing}; "
                         f"headers={hdrs}")
    return cols


def parse_bom(ws, cols: dict[str, int]) -> list[dict]:
    rows = []
    for r in range(2, ws.max_row + 1):
        level = ws.cell(r, cols["level"]).value
        if level is None:
            continue
        rows.append({
            "row": r,
            "level": str(level).strip(),
            "drawing_no": str(ws.cell(r, cols["drawing"]).value or "").strip(),
            "name": str(ws.cell(r, cols["name"]).value or "").strip(),
            "spec": str(ws.cell(r, cols["spec"]).value or "").strip(),
            "quantity": ws.cell(r, cols["qty"]).value,
            "unit": str(ws.cell(r, cols["unit"]).value or "").strip(),
            "step_text": str(ws.cell(r, cols["step"]).value or "").strip(),
        })
    return rows


def sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def norm_key(s: str) -> str:
    """cad-match/v1 key normalisation: case-free, '.' == '_'."""
    return re.sub(r"\s+", "", str(s or "")).lower().replace(".", "_")


def split_level(level: str) -> list[int]:
    return [int(p) for p in level.split(".") if p.strip()]



# ------------------------------------------------------------ CAD matching
def build_stem_index(graph: dict) -> dict[str, list[dict]]:
    """normalised model stem -> occurrences (path order)."""
    index: dict[str, list[dict]] = {}
    for occ in sorted(graph["occurrences"], key=lambda o: o["path"]):
        stem = norm_key(occ["model"].split(".")[0])
        index.setdefault(stem, []).append(occ)
    return index


def allocate_row(row: dict, index: dict, pool_left: dict[str, list[dict]]) \
        -> list[dict]:
    """cad-match/v1: drawing number first, then model/spec; allocate the
    row's quantity from the shared pool in BOM order."""
    if row["unit"] in RAW_UNITS:
        return []                      # material row: no own occurrence
    want = row["quantity"]
    want = int(want) if isinstance(want, (int, float)) and want >= 1 else 1
    for key in (norm_key(row["drawing_no"]), norm_key(row["spec"])):
        if not key:
            continue
        avail = pool_left.get(key, [])
        if avail:
            taken = avail[:want]
            pool_left[key] = avail[want:]
            if taken:
                return taken
    raise SystemExit(
        f"[PLAN] no-silent-skip violated: BOM row {row['row']} "
        f"({row['level']} {row['name']}) matched no CAD occurrence.\n"
        f"       tried keys: drawing_no={row['drawing_no']!r} "
        f"spec={row['spec']!r} (normalised: "
        f"{norm_key(row['drawing_no'])!r}, {norm_key(row['spec'])!r})\n"
        "       fix the BOM key or add the missing CAD model; a step is "
        "never silently skipped.")


# ---------------------------------------------------------------- geometry
def bbox_center(bb: list[float]) -> list[float]:
    return [(bb[0] + bb[3]) / 2, (bb[1] + bb[4]) / 2, (bb[2] + bb[5]) / 2]


def union_bbox(bbs: list[list[float]]) -> list[float] | None:
    bbs = [b for b in bbs if b]
    if not bbs:
        return None
    out = list(bbs[0])
    for b in bbs[1:]:
        for a in range(3):
            out[a] = min(out[a], b[a])
            out[a + 3] = max(out[a + 3], b[a + 3])
    return out


def root_bbox(root: dict, graph: dict) -> list[float]:
    """An assembly root's own bbox can be degenerate in discovery output;
    whenever the root has descendants, the union of the descendant boxes
    is the trustworthy extent."""
    prefix = root["path"] + "/"
    desc = [o["world_bbox"] for o in graph["occurrences"]
            if o["path"].startswith(prefix) and o.get("world_bbox")]
    if desc:
        return union_bbox(desc)
    if root.get("world_bbox"):
        return root["world_bbox"]
    raise SystemExit(f"occurrence {root['path']} has no usable bbox")


def group_roots(occs: list[dict]) -> list[dict]:
    """Drop descendants covered by an ancestor occurrence."""
    roots: list[dict] = []
    for o in sorted(occs, key=lambda x: x["path"]):
        if not any(o["path"].startswith(r["path"] + "/") for r in roots):
            roots.append(o)
    return roots


def derive_explosion(bb: list[float], receiver_bbs: list[list[float]],
                     hull_lo: list[float], hull_hi: list[float],
                     view_from: dict[str, list[float]] | None = None) -> dict:
    """explosion-contact/v4: receiver-face normal from geometry only.

    Receivers are leaf occurrences installed so far.  Contact search runs
    on EVERY world axis (installs are not all vertical).  Insertion mates
    (the part interpenetrates the receiver shallowly along the mate axis,
    e.g. a screw shaft in a hole or a hose over a stub) outrank touch
    contacts (a fastener can brush a side wall without mating with it);
    the interpenetration cap scales with the receiver extent along the
    axis because a shaft may pass clean through a thin receiver.  Without
    any candidate the part exits along the shortest whole-hull clearance
    travel.  The sign always points away from the receiver.  Distance
    stays moderate: proportional to the part extent along the normal,
    clamped.  v5: the v4 fixed 60mm floor exploded tiny seals (a 16mm
    O-ring floated ~4x its own size off its seat and read as a stray
    marker in the image); the floor now scales with the part's own
    extent, and an insertion mate must additionally clear its mate
    depth plus a visual gap.  Large parts keep the 60mm floor.
    """
    diag = math.sqrt(sum((bb[a + 3] - bb[a]) ** 2 for a in range(3)))
    tol_touch = max(2.0, 0.02 * diag)
    base_cap = 0.25 * diag
    touch: list[tuple] = []
    insert: list[tuple] = []
    for rb in receiver_bbs:
        for axis in range(3):
            ov = [min(bb[b + 3], rb[b + 3]) - max(bb[b], rb[b])
                  for b in range(3) if b != axis]
            if min(ov) <= 0:
                continue                       # no facing plane
            area = ov[0] * ov[1]
            lo = rb[axis] - bb[axis + 3]       # receiver on the + side
            hi = bb[axis] - rb[axis + 3]       # part on the + side
            sep, sign = (hi, 1.0) if hi >= lo else (lo, -1.0)
            cap = base_cap + (rb[axis + 3] - rb[axis])
            if sep > tol_touch or sep < -cap:
                continue                       # too far / too deep
            if sep >= -tol_touch:
                touch.append((area, -abs(sep), axis, sign))
            else:
                insert.append((abs(sep), -area, axis, sign))
    depth = 0.0
    if insert:
        depth, _, axis, sign = min(insert)
        method = "insertion_contact"
    elif touch:
        _, _, axis, sign = max(touch)
        method = "face_contact"
    else:
        # shortest hull-clearance travel over all axes and both signs
        best = None
        for axis in range(3):
            for travel, sign in ((hull_hi[axis] - bb[axis + 3], 1.0),
                                 (bb[axis] - hull_lo[axis], -1.0)):
                if best is None or travel < best[0]:
                    best = (travel, axis, sign)
        _, axis, sign = best
        method = "hull_clearance"
    normal = [0.0, 0.0, 0.0]
    normal[axis] = sign
    extent = bb[axis + 3] - bb[axis]
    size = max(bb[a + 3] - bb[a] for a in range(3))
    floor = min(EXP["floor_cap_mm"],
                max(EXP["size_factor"] * size,
                    depth + EXP["insert_extent_frac"] * extent
                    + EXP["insert_gap_mm"]))
    distance = min(EXP["distance_cap_mm"],
                   max(floor, EXP["distance_extent_frac"] * extent))
    # explode-visibility/v1: the away-from-receiver sign is a LOCAL rule;
    # it can drive the part INTO another context solid (observed: a clamp
    # ring pushed below the top plate - fully occluded in the full view
    # while the focus rep still showed it, arrow projecting onto blank
    # plate).  v2: the v1 receiver-AABB interpenetration proxy is blind
    # to CAVITY occlusion (a part inside an enclosed cavity intersects
    # no box yet is invisible from every outside camera), so the primary
    # score is view-ray occlusion: from the exploded centre toward each
    # locked camera eye, count installed boxes crossing the ray; prefer
    # the sign with a clear ray on at least one camera.  v1 hits break
    # ties, then the contact rule's sign.
    centre = [0.5 * (bb[b] + bb[b + 3]) for b in range(3)]

    def _ray_blockers(sgn):
        if not view_from:
            return 0
        c = [centre[b] + (sgn * distance if b == axis else 0.0)
             for b in range(3)]
        per_cam = []
        for vec in view_from.values():
            n = 0
            for rb in receiver_bbs:
                if _ray_hits_box(c, vec, rb):
                    n += 1
            per_cam.append(n)
        return min(per_cam)

    def _exploded_hits(sgn):
        shift = sgn * distance
        n = 0
        for rb in receiver_bbs:
            if all(min(bb[b + 3] + (shift if b == axis else 0.0), rb[b + 3])
                   - max(bb[b] + (shift if b == axis else 0.0), rb[b]) > 0.0
                   for b in range(3)):
                n += 1
        return n

    def _score(sgn):
        return (_ray_blockers(sgn), _exploded_hits(sgn))
    if _score(-sign) < _score(sign):
        sign = -sign
        normal[axis] = sign
        method += "_visibility_flip"
    return {"normal": normal, "distance": distance, "axis": axis,
            "method": method}


def _ray_hits_box(c: list[float], d: list[float], rb: list[float],
                  length: float = 1e4, eps: float = 1e-9) -> bool:
    """Segment c->c+d*length vs AABB slab test; boxes containing the ray
    origin (interpenetration, v1's territory) are NOT occluders."""
    if all(rb[b] <= c[b] <= rb[b + 3] for b in range(3)):
        return False
    tmin, tmax = 0.0, length
    for b in range(3):
        if abs(d[b]) < eps:
            if c[b] < rb[b] or c[b] > rb[b + 3]:
                return False
        else:
            t1 = (rb[b] - c[b]) / d[b]
            t2 = (rb[b + 3] - c[b]) / d[b]
            if t1 > t2:
                t1, t2 = t2, t1
            tmin = max(tmin, t1)
            tmax = min(tmax, t2)
            if tmin > tmax:
                return False
    return tmax > tmin


# ----------------------------------------------------------------- cameras
def camera_vectors(manifest: dict) -> dict[str, list[float]] | None:
    """Unit view-FROM direction per camera, from the calibrated manifest.
    The stored matrix is transposed (Creo applies M^T), so view rows are
    stored columns; stored column 2 IS the view-from direction (verified
    against rendered images: fixed_123 looks into the +Y opening)."""
    cams = manifest.get("cameras")
    if not cams:
        return None
    out = {}
    for cid, mat in cams.items():
        vfrom = [mat[k][2] for k in range(3)]
        n = math.sqrt(sum(v * v for v in vfrom))
        out[cid] = [v / n for v in vfrom]
    return out


# explosion-screen-floor/v1: whole-machine finals (sop-context/v1)
# shrink every millimetre to ~1 px, so the v5 hug-the-seat floor can
# leave the exploded part visually fused with its seat (30.22/23/25
# V3).  Enforce a minimum ON-SCREEN separation on the view plane.
SCREEN_FLOOR_PX = EXP["screen_floor_px"]
SCREEN_FLOOR_MAX_SCALE = EXP["screen_floor_max_scale"]
# mirrors the Renderer framing policy (image 1600, margins, occupancy)
_FRM = _CFG["framing_mirror"]
_AVAIL_W = _FRM["image_px"] - _FRM["margin_l_px"] - _FRM["margin_r_px"]
_AVAIL_H = _FRM["image_px"] - _FRM["margin_t_px"] - _FRM["margin_b_px"]
_OCC = _FRM["occupancy"]


def camera_basis(manifest: dict) -> dict[str, tuple[list[float], list[float]]]:
    """Unit right/up view rows per camera (stored columns 0/1 of the
    transposed matrix, same convention as camera_vectors)."""
    cams = manifest.get("cameras") or {}
    out = {}
    for cid, mat in cams.items():
        basis = []
        for col in (0, 1):
            v = [mat[k][col] for k in range(3)]
            n = math.sqrt(sum(x * x for x in v)) or 1.0
            basis.append([x / n for x in v])
        out[cid] = (basis[0], basis[1])
    return out


def estimate_px_per_mm(basis: tuple[list[float], list[float]],
                       hull_lo: list[float], hull_hi: list[float]) -> float:
    """Closed-form framing fit of the hull on the view plane: the px
    per mm the Renderer's zoom solve will land near (occupancy fit)."""
    right, up = basis
    lo = math.inf
    hi_w = hi_h = -math.inf
    lo_w = lo_h = math.inf
    for x in (hull_lo[0], hull_hi[0]):
        for y in (hull_lo[1], hull_hi[1]):
            for z in (hull_lo[2], hull_hi[2]):
                w = right[0] * x + right[1] * y + right[2] * z
                h = up[0] * x + up[1] * y + up[2] * z
                lo_w = min(lo_w, w)
                hi_w = max(hi_w, w)
                lo_h = min(lo_h, h)
                hi_h = max(hi_h, h)
    w_mm = max(hi_w - lo_w, 1.0)
    h_mm = max(hi_h - lo_h, 1.0)
    return min(_OCC * _AVAIL_W / w_mm, _OCC * _AVAIL_H / h_mm)


def choose_camera(normal: list[float], exploded_center: list[float],
                  receiver_center: list[float] | None,
                  view_from: dict[str, list[float]] | None) -> tuple[str, str]:
    """camera-approach/v3: prefer the camera whose eye stands on the side
    the part comes from (behind the moving part, so the receiver face is
    visible); occlusion tie-break; analytic mirror fallback before
    calibration exists."""
    if not view_from:
        cam = "fixed_123" if normal[1] > 0 else "fixed_456"
        return cam, ("analytic_mirror_fallback (no calibrated cameras "
                     "yet; vertical-mirror pair, Y component decides)")
    scores = {cid: sum(f * n for f, n in zip(vec, normal))
              for cid, vec in view_from.items()}
    ranked = sorted(scores.items(), key=lambda kv: -kv[1])
    (ca, sa), (cb, sb) = ranked[0], ranked[1]
    if sa > sb + 1e-6:
        return ca, f"approach_side score {sa:.4f} vs {sb:.4f}"
    if sb > sa + 1e-6:
        return cb, f"approach_side score {sb:.4f} vs {sa:.4f}"
    if receiver_center is not None:
        rel = [e - r for e, r in zip(exploded_center, receiver_center)]
        tie = {cid: sum(f * v for f, v in zip(vec, rel))
               for cid, vec in view_from.items()}
        ranked = sorted(tie.items(), key=lambda kv: -kv[1])
        (ca, ta), (cb, tb) = ranked[0], ranked[1]
        if ta > tb + 1e-6:
            return ca, f"occlusion_tiebreak {ta:.1f} vs {tb:.1f}"
        if tb > ta + 1e-6:
            return cb, f"occlusion_tiebreak {tb:.1f} vs {ta:.1f}"
    return "fixed_123", "approach scores equal, default view"


# -------------------------------------------------------------------- main
def main() -> None:
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    flags = {a for a in sys.argv[1:] if a.startswith("--")}
    if not args or not re.fullmatch(r"[a-z0-9_]{3,40}", args[0]):
        raise SystemExit("usage: plan_steps.py <task_code> [limit] "
                         "[--reuse-batch]   (task_code: [a-z0-9_]{3,40})")
    task_code = args[0]
    limit = int(args[1]) if len(args) > 1 else 10 ** 9

    graph = json.loads(GRAPH_FILE.read_text(encoding="utf-8"))
    bom_ws, bom_headers = open_bom_sheet()
    bom_rows = parse_bom(bom_ws, detect_columns(bom_headers))
    occ_by_path = {o["path"]: o for o in graph["occurrences"]}
    # leaf occurrences = no child paths in the graph; only leaves are
    # trustworthy explosion receivers (explosion-contact/v4).
    has_child = {o["path"].rsplit("/", 1)[0] for o in graph["occurrences"]}

    # ---------- batch-scoped image folder (naming/v1) ----------
    images_dir = None
    if "--reuse-batch" in flags and OUT_PLAN.exists():
        old = json.loads(OUT_PLAN.read_text(encoding="utf-8"))
        if old.get("task_code") == task_code:
            images_dir = old.get("images_dir")
    if not images_dir:
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M")
        images_dir = f"out/images/{task_code}_{stamp}"

    # ---------- manifest ----------
    asm_path = MODELS_DIR / ASM_FILE
    manifest = {
        "schema": "clean-run-manifest/v1",
        "final_assembly_file": ASM_FILE,
        "final_assembly_version": graph.get("assembly_version"),
        "sha256_at_batch_start": sha256_of(asm_path),
        "root_token": graph.get("root_token"),
        "cameras_note": ("cameras are added by calibrate_cameras.py "
                         "(dual view: top + mirrored bottom isometric)"),
    }
    if OUT_MANIFEST.exists():
        # keep already-calibrated cameras across replans
        prev = json.loads(OUT_MANIFEST.read_text(encoding="utf-8"))
        for key in ("cameras", "camera_note", "screen_cal"):
            if key in prev:
                manifest[key] = prev[key]
    OUT_MANIFEST.write_text(json.dumps(manifest, ensure_ascii=False, indent=2),
                            encoding="utf-8")
    view_from = camera_vectors(manifest)
    cam_basis = camera_basis(manifest)

    # ---------- main steps from the process column (bom-steps/v2) ----
    # "第N步…" anchors start main assembly steps; every following row
    # rides the current main step.  Rows before the first anchor (the
    # product root / inspection notes) never gate rendering.
    root_depth = min(len(split_level(r["level"])) for r in bom_rows)
    main_steps: list[dict] = []
    cur: dict | None = None
    for row in bom_rows:
        no = parse_step_no(row["step_text"])
        if no is not None:
            cur = {"no": no, "anchor": row, "rows": []}
            main_steps.append(cur)
        if cur is None:
            if len(split_level(row["level"])) <= root_depth:
                continue                   # product root before any anchor
            cur = {"no": 0, "anchor": row, "rows": []}   # untagged rows
            main_steps.append(cur)
        cur["rows"].append(row)

    # ---------- render sub-step units inside each main step ----------
    # direct-child groups of the product root, created on first sight so
    # a subtree like 30.20.x with no own anchor row still gets its unit.
    # first-placement/v2: the first main step that is a single assembly
    # with listed children is the base build - the anchor becomes a
    # container (no own occurrence) and its direct children are units.
    direct_depth = root_depth + 1
    units: list[dict] = []
    container_rows: list[dict] = [
        r for r in bom_rows
        if len(split_level(r["level"])) == root_depth]
    expanded_once = False
    for ms in main_steps:
        groups: list[dict] = []
        by_key: dict[tuple, dict] = {}
        for row in ms["rows"]:
            parts = split_level(row["level"])
            if len(parts) <= root_depth:
                continue                   # product root row
            key = tuple(parts[:direct_depth])
            g = by_key.get(key)
            if g is None:
                g = {"level": ".".join(str(p) for p in parts[:direct_depth]),
                     "anchor": row, "rows": [row], "material_rows": [],
                     "main_no": ms["no"]}
                by_key[key] = g
                groups.append(g)
            else:
                g["rows"].append(row)
        if not groups:
            continue                       # inspection-style main step
        ms_title = re.sub(r"\s+", " ", ms["anchor"]["step_text"]).strip()
        base = groups[0]
        has_deeper = any(len(split_level(r["level"])) > direct_depth
                         for r in base["rows"])
        if not expanded_once and len(groups) == 1 and has_deeper:
            expanded_once = True           # base structure built in place
            container_rows.append(base["anchor"])
            cdepth = len(split_level(base["anchor"]["level"]))
            cby: dict[tuple, dict] = {}
            for row in base["rows"]:
                if row is base["anchor"]:
                    continue               # container, rides no step
                parts = split_level(row["level"])
                key = tuple(parts[:cdepth + 1])
                cg = cby.get(key)
                if cg is None:
                    cg = {"level": ".".join(str(p) for p in parts[:cdepth + 1]),
                          "anchor": row, "rows": [row], "material_rows": [],
                          "main_no": ms["no"], "main_title": ms_title}
                    cby[key] = cg
                    units.append(cg)
                else:
                    cg["rows"].append(row)
        else:
            for g in groups:
                g["main_title"] = ms_title
                units.append(g)

    per_main: dict[int, int] = {}
    for u in units:
        per_main[u["main_no"]] = per_main.get(u["main_no"], 0) + 1
        u["sub_index"] = per_main[u["main_no"]]
    main_w = max(2, len(str(max((u["main_no"] for u in units), default=1))))
    sub_w = max(2, len(str(max((u["sub_index"] for u in units), default=1))))
    main_step_meta = [
        {"no": ms["no"],
         "title": re.sub(r"\s+", " ", ms["anchor"]["step_text"]).strip(),
         "anchor_level": ms["anchor"]["level"]}
        for ms in main_steps]
    # container stems are excluded from the no-silent-skip audit below
    container_keys = set()
    for r in container_rows:
        container_keys.add(norm_key(r["drawing_no"]))
        container_keys.add(norm_key(r["spec"]))

    # ---------- multi-key CAD allocation (cad-match/v1) ----------
    index = build_stem_index(graph)
    pool_left = {k: list(v) for k, v in index.items()}
    for g in units:
        occs: list[dict] = []
        for row in g["rows"]:
            if row["unit"] in RAW_UNITS:
                g["material_rows"].append(
                    f"{row['level']} {row['name']} "
                    f"{row['quantity']}{row['unit']} (raw material, "
                    "rides with its fabricated parent)")
                continue
            occs.extend(allocate_row(row, index, pool_left))
        g["occs"] = occs

    # whole-assembly hull per axis (every occurrence with a bbox)
    all_bbs = [o["world_bbox"] for o in graph["occurrences"]
               if o.get("world_bbox")]
    hull_lo = [min(b[a] for b in all_bbs) for a in range(3)]
    hull_hi = [max(b[a + 3] for b in all_bbs) for a in range(3)]

    # ---------- build renderable steps in strict BOM order ----------
    steps: list[dict] = []
    merged_rows: list[str] = []
    completed_paths: list[str] = []
    placed_paths: list[str] = []           # every path installed so far
    receiver_bbs: list[list[float]] = []   # installed parts' boxes
    receiver_centers: list[list[float]] = []

    for g in units:
        occs = g["occs"]
        roots = group_roots(occs)
        # component-detail rule: a root buried inside an already-installed
        # occurrence was installed with it; record it explicitly instead
        # of exploding an installed part.
        kept: list[dict] = []
        for r in roots:
            owner = next((p for p in placed_paths
                          if r["path"].startswith(p + "/")), None)
            if owner is not None:
                row_txt = next((row["level"] for row in g["rows"]
                                if norm_key(row["drawing_no"]) ==
                                norm_key(r["model"].split(".")[0]) or
                                norm_key(row["spec"]) ==
                                norm_key(r["model"].split(".")[0])),
                               "?")
                merged_rows.append(
                    f"{r['path']} ({r['model']}, BOM row {row_txt}): "
                    f"component-detail of group {g['level']}; installed "
                    f"riding inside {owner}")
            else:
                kept.append(r)
        if not kept:
            continue                       # whole group rode an ancestor
        roots = kept

        first_step = not steps
        step_occs_paths = sorted(o["path"] for o in occs)
        # the visible set must list every leaf that moves or stays: use
        # all matched occurrences of the group plus, for each kept root,
        # every graph descendant (matched or not) so no rider is lost.
        rider_paths = set(step_occs_paths)
        for r in roots:
            prefix = r["path"] + "/"
            rider_paths |= {o["path"] for o in graph["occurrences"]
                            if o["path"].startswith(prefix)}
        # rep chain (bom-steps/v2): every ancestor occurrence of a
        # visible leaf must be INCLUDED too, else the SimpRep default
        # EXCLUDE prunes the subtree above it (container assemblies
        # like the base weldment ride no unit of their own)
        for p in list(rider_paths):
            parts = p.split("/")
            for i in range(1, len(parts)):
                a = "/".join(parts[:i])
                if a in occ_by_path:
                    rider_paths.add(a)
        step_paths = sorted(rider_paths)

        moving = []
        if first_step:
            # first-placement/v1: the base weldment lands as one rigid
            # unit; there is no receiver yet, so the separation is the
            # vertical placement direction (+Y exploded = lowered down).
            gbb = union_bbox([root_bbox(r, graph) for r in roots])
            distance = min(EXP["distance_cap_mm"],
                           max(EXP["fallback_floor_mm"],
                               EXP["distance_extent_frac"]
                               * (gbb[4] - gbb[1])))
            translation = [0.0, distance, 0.0]
            center = bbox_center(gbb)
            # bottom-face centre: the projected arrow then runs from the
            # hovering weldment down to the (annotated) fixture level.
            anchor = [center[0], gbb[1], center[2]]
            # EVERY root of the weldment lands rigidly: a weldment can
            # span sibling subtrees, and a root missing from `moving`
            # leaves its whole subtree parked mid-air.
            for r in roots:
                moving.append({
                    "path": r["path"],
                    "model": r["model"],
                    "translation": translation,
                    "explode_distance": round(distance, 3),
                    "explode_axis": 1,
                    "explode_method": "first_placement",
                    "anchor_complete": anchor,
                    "anchor_exploded": [anchor[i] + translation[i]
                                        for i in range(3)],
                })
            evidence = {"method": "first_placement",
                        "note": "base weldment placed onto the fixture; "
                                "no receiver geometry exists yet"}
            camera, cam_why = "fixed_123", "first placement, default view"
            exploded_center = [center[0], center[1] + distance, center[2]]
        else:
            evidence_notes = []
            exploded_center = None
            for r in roots:
                bb = root_bbox(r, graph)
                d = derive_explosion(bb, receiver_bbs, hull_lo, hull_hi,
                                     view_from)
                normal = d["normal"]
                distance = d["distance"]
                translation = [n * distance for n in normal]
                center = bbox_center(bb)
                anchor = [
                    bb[0] if normal[0] < 0 else
                    (bb[3] if normal[0] > 0 else center[0]),
                    bb[1] if normal[1] < 0 else
                    (bb[4] if normal[1] > 0 else center[1]),
                    bb[2] if normal[2] < 0 else
                    (bb[5] if normal[2] > 0 else center[2]),
                ]
                moving.append({
                    "path": r["path"],
                    "model": r["model"],
                    "translation": translation,
                    "explode_distance": round(distance, 3),
                    "explode_axis": d["axis"],
                    "explode_method": d["method"],
                    "anchor_complete": anchor,
                    "anchor_exploded": [anchor[i] + translation[i]
                                        for i in range(3)],
                })
                evidence_notes.append(
                    f"{r['path']}: {d['method']} axis={d['axis']} "
                    f"sep_normal={normal}")
                if exploded_center is None:
                    exploded_center = [center[i] + translation[i]
                                       for i in range(3)]
            receiver_center = (
                [sum(v) / len(receiver_centers)
                 for v in zip(*receiver_centers)]
                if receiver_centers else None)
            primary = moving[0]["translation"]
            nrm = math.sqrt(sum(v * v for v in primary)) or 1.0
            camera, cam_why = choose_camera(
                [v / nrm for v in primary], exploded_center,
                receiver_center, view_from)

            # explosion-screen-floor/v1: raise distances whose on-screen
            # view-plane separation falls under SCREEN_FLOOR_PX at the
            # estimated final scale; direction stays exact, only the
            # magnitude grows (capped).
            if camera in cam_basis:
                k_est = estimate_px_per_mm(cam_basis[camera],
                                           hull_lo, hull_hi)
                for m in moving:
                    tr = m["translation"]
                    tl = math.sqrt(sum(v * v for v in tr))
                    if tl < 1e-9:
                        continue
                    uw = sum(cam_basis[camera][0][i] * tr[i]
                             for i in range(3)) / tl
                    uh = sum(cam_basis[camera][1][i] * tr[i]
                             for i in range(3)) / tl
                    sep_unit = max(math.hypot(uw, uh), 0.25)
                    need = SCREEN_FLOOR_PX / (k_est * sep_unit)
                    if need > tl:
                        scale = min(need / tl, SCREEN_FLOOR_MAX_SCALE)
                        tr = [v * scale for v in tr]
                        m["translation"] = tr
                        m["explode_distance"] = round(tl * scale, 3)
                        m["screen_floor_scale"] = round(scale, 2)
                        m["anchor_exploded"] = [
                            m["anchor_complete"][i] + tr[i]
                            for i in range(3)]
            evidence = {"receiver_center": receiver_center,
                        "notes": evidence_notes}

        # rigid-group riders: descendants ride the ancestor translation,
        # one representative arrow per group (render-rules: Arrows).
        for o in occs:
            if any(o["path"] == r["path"] for r in roots):
                continue
            root = next((r for r in roots
                         if o["path"].startswith(r["path"] + "/")), None)
            if root is None:
                continue                   # sibling of a merged root
            anchor = bbox_center(o.get("world_bbox") or [0.0] * 6)
            moving.append({
                "path": o["path"],
                "model": o["model"],
                "translation": [0.0, 0.0, 0.0],
                "rides_with": root["path"],
                "anchor_complete": anchor,
                "anchor_exploded": anchor,
            })

        raw_title = g["anchor"]["step_text"] or f"安装{g['anchor']['name']}"
        title = re.sub(r"\s+", " ", raw_title).strip()
        steps.append({
            "schema": "clean-run-step/v3",
            "step_id": (f"{str(g['main_no']).zfill(main_w)}."
                        f"{str(g['sub_index']).zfill(sub_w)}_{task_code}"),
            "main_step": g["main_no"],
            "main_step_title": g.get("main_title", ""),
            "sub_index": g["sub_index"],
            "bom_level": g["level"],
            "title": title,
            "bom_names": [r["name"] for r in g["rows"]],
            "bom_quantities": [r["quantity"] for r in g["rows"]],
            "material_rows": g["material_rows"],
            "moving": moving,
            "moving_paths": step_paths,
            "receiver_paths": sorted(set(completed_paths)),
            "visible_paths": sorted(set(completed_paths + step_paths)),
            "camera": camera,
            "camera_evidence": cam_why,
            "explosion_policy": "contact-normal/v5",
            "explosion_evidence": evidence,
        })
        completed_paths += step_paths
        placed_paths += [r["path"] for r in roots]
        for r in roots:
            receiver_centers.append(bbox_center(root_bbox(r, graph)))
        # leaf-level receivers for the NEXT steps' contact search
        for p in step_paths:
            o = occ_by_path.get(p)
            if o and o.get("world_bbox") and p not in has_child:
                receiver_bbs.append(o["world_bbox"])
        if len(steps) >= limit:
            break

    # ---------- audit trail: nothing may vanish silently ----------
    unallocated = {k: [o["path"] for o in v]
                   for k, v in pool_left.items()
                   if v and k not in container_keys}
    plan = {
        "schema": "clean-run-plan/v3",
        "task_code": task_code,
        "images_dir": images_dir,
        "manifest": str(OUT_MANIFEST.relative_to(ROOT)),
        "assembly": graph["assembly_file"],
        "occurrence_count_total": graph["occurrence_count"],
        "main_steps": main_step_meta,
        "container_rows": [f"{r['level']} {r['name']} "
                           f"{r['drawing_no']} (container, rides no step)"
                           for r in container_rows],
        "steps": steps,
        "merged_component_detail_rows": merged_rows,
        "unallocated_occurrences": unallocated,
        "rules": ["bom-steps/v2", "bom-order-authority/v1",
                  "no-silent-skip/v1", "cad-match/v1",
                  "first-placement/v2", "explosion-contact/v5",
                  "camera-approach/v3", "naming/v2"],
    }
    OUT_PLAN.write_text(json.dumps(plan, ensure_ascii=False, indent=2),
                        encoding="utf-8")

    print(f"[PLAN] task_code={task_code} images_dir={images_dir}")
    for ms in main_step_meta:
        subs = [s for s in steps if s["main_step"] == ms["no"]]
        print(f"[PLAN] main step {ms['no']}: {ms['title'][:46]!r} "
              f"anchor={ms['anchor_level']} sub_steps={len(subs)}")
        for s in subs:
            roots_n = sum(1 for m in s["moving"] if "rides_with" not in m)
            print(f"[PLAN]   {s['step_id']} (bom {s['bom_level']}): "
                  f"{s['title'][:30]!r} roots={roots_n} "
                  f"moving={len(s['moving'])} "
                  f"visible={len(s['visible_paths'])} cam={s['camera']}")
    for m in merged_rows:
        print("[PLAN] merged component-detail:", m)
    for m in sum((g["material_rows"] for g in units), []):
        print("[PLAN] material row:", m)
    for c in plan["container_rows"]:
        print("[PLAN] container row:", c)
    if unallocated:
        print("[PLAN] occurrences with no BOM row (stay unrendered):")
        for k, v in unallocated.items():
            print("        -", k, v)
    print(f"[PLAN] wrote {OUT_PLAN}")


if __name__ == "__main__":
    main()
